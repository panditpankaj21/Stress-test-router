"""
Auto-updating Excel Statistics Generator with Detailed Failure Analysis
Creates a comprehensive test history with clickable navigation, visual progress bars, and step-level debugging
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import os
from collections import defaultdict


class ExcelStatisticsGenerator:
    def __init__(self, excel_path: str = "test_reports/test_statistics.xlsx"):
        """Initialize Excel statistics generator"""
        self.excel_path = excel_path
        self.ensure_excel_exists()
        
        # Define color schemes
        self.colors = {
            'header': '366092',
            'pass': '92D050',      # Green
            'fail': 'FF0000',      # Red
            'warning': 'FFC000',   # Orange/Yellow
            'subheader': 'D9E1F2',
            'link': '0563C1',
            'border': '000000',
            'step_passed': 'C6EFCE',   # Light green
            'step_failed': 'FFC7CE',   # Light red
            'step_skipped': 'FFEB9C',  # Light yellow
            'step_undefined': 'E7E6E6' # Light gray
        }
        
        # Define styles
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Calibri', size=11, bold=True)
        self.normal_font = Font(name='Calibri', size=10)
        self.bold_font = Font(name='Calibri', size=10, bold=True)
        self.link_font = Font(name='Calibri', size=10, color=self.colors['link'], underline='single')
        self.small_font = Font(name='Calibri', size=8)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def ensure_excel_exists(self):
        """Ensure the Excel file and directory exist"""
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            wb.save(self.excel_path)
            wb.close()
    
    def _get_router_sheet_name(self, router_name: str, router_model: str) -> str:
        """Generate safe sheet name: Name-Model"""
        sheet_name = f"{router_name}-{router_model}"
        sheet_name = sheet_name[:31]
        for char in ['/', '\\', '?', '*', '[', ']', ':']:
            sheet_name = sheet_name.replace(char, '_')
        return sheet_name
    
    def _calculate_duration(self, start_time, end_time) -> str:
        """Calculate duration between start and end time and format as HH:MM:SS"""
        try:
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            if isinstance(end_time, str):
                end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            
            duration = end_time - start_time
            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        except:
            return "00:00:00"
    
    def _aggregate_router_data(self, all_tests: List[Dict]) -> Dict:
        """Aggregate test data by router"""
        aggregated = defaultdict(lambda: {
            'router_name': None,
            'router_model': None,
            'router_firmware': None,
            'router_mac': None,
            'total_tests': 0,
            'passed_tests': 0,
            'failed_tests': 0,
            'total_duration': 0,
            'features': defaultdict(list),
            'failed_tests_details': []  # NEW: Track failed tests for analysis
        })
        
        for test in all_tests:
            router_mac = test.get('router_mac')
            if not router_mac:
                continue
            
            router_data = aggregated[router_mac]
            
            if not router_data['router_name']:
                router_data['router_name'] = test.get('router_name', 'Unknown')
                router_data['router_model'] = test.get('router_model', 'Unknown')
                router_data['router_firmware'] = test.get('router_firmware', 'Unknown')
                router_data['router_mac'] = router_mac
            
            router_data['total_tests'] += 1
            status = test.get('status', '').lower()
            if status == 'passed':
                router_data['passed_tests'] += 1
            else:
                router_data['failed_tests'] += 1
                # Store failed test for detailed analysis
                router_data['failed_tests_details'].append(test)
            
            # Calculate test duration
            start_time = test.get('start_time')
            end_time = test.get('end_time')
            if start_time and end_time:
                try:
                    if isinstance(start_time, str):
                        start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                    if isinstance(end_time, str):
                        end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
                    
                    duration = (end_time - start_time).total_seconds()
                    router_data['total_duration'] += duration
                except:
                    pass
            
            feature_name = test.get('feature_name', 'Unknown Feature')
            router_data['features'][feature_name].append(test)
        
        return dict(aggregated)
    
    def _format_duration_from_seconds(self, total_seconds: float) -> str:
        """Format total seconds as HH:MM:SS"""
        total_seconds = int(total_seconds)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    def _create_progress_bar_cells(self, ws, row: int, col: int, 
                                   passed: int, total: int, bar_width: int = 20):
        """Create a visual progress bar using colored cells"""
        if total == 0:
            pass_percentage = 0
        else:
            pass_percentage = (passed / total) * 100
        
        green_cells = int((passed / total) * bar_width) if total > 0 else 0
        
        for i in range(bar_width):
            cell = ws.cell(row=row, column=col + i)
            cell.border = self.thin_border
            
            if i < green_cells:
                cell.fill = PatternFill(start_color=self.colors['pass'], 
                                       end_color=self.colors['pass'], 
                                       fill_type='solid')
            else:
                cell.fill = PatternFill(start_color=self.colors['fail'], 
                                       end_color=self.colors['fail'], 
                                       fill_type='solid')
        
        percentage_cell = ws.cell(row=row, column=col + bar_width)
        percentage_cell.value = f"{pass_percentage:.1f}%"
        percentage_cell.font = self.bold_font
        percentage_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _create_summary_sheet(self, wb: Workbook, aggregated_data: Dict):
        """Create or update the summary sheet"""
        
        if "Test Summary" in wb.sheetnames:
            wb.remove(wb["Test Summary"])
        
        ws = wb.create_sheet("Test Summary", 0)
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = '📊 Router Test Statistics - Summary'
        ws['A1'].font = self.header_font
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], 
                                    end_color=self.colors['header'], 
                                    fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Last updated
        row = 3
        ws[f'A{row}'] = 'Last Updated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = self.bold_font
        ws.merge_cells(f'B{row}:C{row}')
        
        # Add link to failure analysis
        row += 1
        ws[f'A{row}'] = '🔍 View Detailed Failure Analysis →'
        ws[f'A{row}'].hyperlink = "#'Failure Analysis'!A1"
        ws[f'A{row}'].font = self.link_font
        
        # Headers for router list
        row += 2
        headers = ['Router Name', 'Tests (Passed/Total)', 'Total Duration', 'Success Rate']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                   end_color=self.colors['header'], 
                                   fill_type='solid')
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.row_dimensions[row].height = 25
        
        # Router data rows
        row += 1
        for router_mac, router_data in sorted(aggregated_data.items(), 
                                             key=lambda x: x[1]['router_name']):
            router_name = router_data['router_name']
            router_model = router_data['router_model']
            total_tests = router_data['total_tests']
            passed_tests = router_data['passed_tests']
            total_duration = router_data['total_duration']
            
            sheet_name = self._get_router_sheet_name(router_name, router_model)
            router_display = f"{router_name}-{router_model}"
            
            cell = ws.cell(row=row, column=1)
            cell.value = router_display
            cell.hyperlink = f"#{sheet_name}!A1"
            cell.font = self.link_font
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            cell = ws.cell(row=row, column=2)
            cell.value = f"{passed_tests}/{total_tests}"
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.normal_font
            
            cell = ws.cell(row=row, column=3)
            cell.value = self._format_duration_from_seconds(total_duration)
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.normal_font
            
            self._create_progress_bar_cells(ws, row, 4, passed_tests, total_tests, bar_width=20)
            
            row += 1
        
        # Legend
        row += 2
        ws[f'A{row}'] = 'Legend:'
        ws[f'A{row}'].font = self.bold_font
        
        row += 1
        ws[f'A{row}'] = 'Green'
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['pass'], 
                                        end_color=self.colors['pass'], 
                                        fill_type='solid')
        ws[f'A{row}'].border = self.thin_border
        ws[f'B{row}'] = 'Passed Tests'
        
        row += 1
        ws[f'A{row}'] = 'Red'
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['fail'], 
                                        end_color=self.colors['fail'], 
                                        fill_type='solid')
        ws[f'A{row}'].border = self.thin_border
        ws[f'B{row}'] = 'Failed Tests'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        for i in range(4, 25):
            ws.column_dimensions[get_column_letter(i)].width = 2
    
    def _create_failure_analysis_sheet(self, wb: Workbook, aggregated_data: Dict):
        """Create a dedicated sheet for detailed failure analysis with enhanced null handling"""
        
        if "Failure Analysis" in wb.sheetnames:
            wb.remove(wb["Failure Analysis"])
        
        ws = wb.create_sheet("Failure Analysis")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = '🔍 Detailed Failure Analysis'
        ws['A1'].font = self.header_font
        ws['A1'].fill = PatternFill(start_color=self.colors['fail'], 
                                    end_color=self.colors['fail'], 
                                    fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Back to summary link
        row = 3
        ws[f'A{row}'] = '← Back to Summary'
        ws[f'A{row}'].hyperlink = "#'Test Summary'!A1"
        ws[f'A{row}'].font = self.link_font
        
        row += 2
        
        # Collect all failed tests
        all_failed_tests = []
        for router_mac, router_data in aggregated_data.items():
            for test in router_data['failed_tests_details']:
                test['_router_name'] = router_data['router_name']
                test['_router_model'] = router_data['router_model']
                all_failed_tests.append(test)
        
        if not all_failed_tests:
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = '🎉 No failed tests found! All tests passed successfully.'
            ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='00B050')
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            return
        
        # Statistics summary
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f'Total Failed Tests: {len(all_failed_tests)}'
        ws[f'A{row}'].font = self.bold_font
        ws[f'A{row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws[f'A{row}'].border = self.thin_border
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Failure type breakdown
        failure_types = {
            'undefined': 0,
            'failed': 0,
            'error': 0
        }
        
        for test in all_failed_tests:
            steps_data = test.get('steps_data', [])
            for step in steps_data:
                if step.get('status') == 'undefined':
                    failure_types['undefined'] += 1
                    break
                elif step.get('status') == 'failed':
                    failure_types['failed'] += 1
                    break
        
        col = 3
        for fail_type, count in failure_types.items():
            if count > 0:
                ws[f'{get_column_letter(col)}{row}'] = f'{fail_type.title()}: {count}'
                ws[f'{get_column_letter(col)}{row}'].font = self.normal_font
                ws[f'{get_column_letter(col)}{row}'].fill = PatternFill(
                    start_color='FFEB9C' if fail_type == 'undefined' else 'FFC7CE',
                    end_color='FFEB9C' if fail_type == 'undefined' else 'FFC7CE',
                    fill_type='solid'
                )
                ws[f'{get_column_letter(col)}{row}'].border = self.thin_border
                ws[f'{get_column_letter(col)}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                col += 1
        
        row += 2
        
        # Headers
        headers = [
            'Router', 'Feature', 'Scenario', 'Date', 'Failure Type',
            'Failed Step', 'Step Status', 'Error Message'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.header_font
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                end_color=self.colors['header'], 
                                fill_type='solid')
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Sort by date (newest first)
        all_failed_tests.sort(key=lambda x: x.get('test_time', ''), reverse=True)
        
        for test in all_failed_tests:
            # Parse test time
            test_time = test.get('test_time', '')
            if test_time:
                if isinstance(test_time, str):
                    try:
                        dt = datetime.fromisoformat(test_time.replace('Z', '+00:00'))
                        test_date = dt.strftime('%d %b %Y %H:%M')
                    except:
                        test_date = test_time[:16] if len(test_time) >= 16 else 'N/A'
                else:
                    test_date = test_time.strftime('%d %b %Y %H:%M')
            else:
                test_date = 'N/A'
            
            # Analyze failure
            steps_data = test.get('steps_data', [])
            failed_step = None
            step_status = None
            error_message = test.get('failure_reason', 'Unknown error')
            failure_type = 'Unknown'
            
            for step in steps_data:
                step_status_val = step.get('status', '')
                if step_status_val in ['failed', 'undefined', 'error']:
                    failed_step = f"{step.get('keyword', '')} {step.get('name', '')}"
                    step_status = step_status_val.upper()
                    error_message = step.get('failure_message', error_message)
                    
                    # Determine failure type
                    if step_status_val == 'undefined':
                        failure_type = 'Missing Step Definition'
                    elif 'timeout' in error_message.lower():
                        failure_type = '⏱ Timeout'
                    elif 'assertion' in error_message.lower():
                        failure_type = 'Assertion Failed'
                    elif 'connection' in error_message.lower() or 'network' in error_message.lower():
                        failure_type = 'Network Error'
                    else:
                        failure_type = 'Runtime Error'
                    break
            
            if not failed_step:
                failed_step = test.get('scenario_name', 'Unknown')
                step_status = 'FAILED'
                failure_type = '⚠️ Test Failed'
            
            # Write data
            data_row = [
                f"{test['_router_name']}-{test['_router_model']}",
                test.get('feature_name', 'N/A'),
                test.get('scenario_name', 'N/A'),
                test_date,
                failure_type,
                failed_step,
                step_status,
                error_message
            ]
            
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                cell.border = self.thin_border
                cell.alignment = Alignment(
                    horizontal='left' if col_idx in [6, 8] else 'center', 
                    vertical='center', 
                    wrap_text=True
                )
                cell.font = Font(name='Calibri', size=9)
            
            # Color code based on failure type
            color = self.colors['step_failed']  # Default red
            if 'Missing Step Definition' in failure_type:
                color = self.colors['step_skipped']  # Yellow for undefined
            elif 'Timeout' in failure_type:
                color = 'FFB366'  # Orange for timeout
            
            # Highlight entire row
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Make step status column bold
            ws.cell(row=row, column=7).font = Font(name='Calibri', size=9, bold=True)
            
            ws.row_dimensions[row].height = 40
            row += 1
        
        # Add legend at bottom
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = 'Failure Type Legend:'
        ws[f'A{row}'].font = self.bold_font
        
        row += 1
        legends = [
            ('Missing Step Definition', 'FFEB9C', 'Step definition not implemented'),
            ('Timeout', 'FFB366', 'Operation exceeded time limit'),
            ('Assertion Failed', 'FFC7CE', 'Test assertion did not pass'),
            ('Network Error', 'FFC7CE', 'Connection or network issue'),
            ('Runtime Error', 'FFC7CE', 'Code execution error'),
            ('Test Failed', 'FFC7CE', 'General test failure')
        ]
        
        for legend_text, color, description in legends:
            ws[f'A{row}'] = legend_text
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].border = self.thin_border
            ws[f'A{row}'].font = Font(name='Calibri', size=9, bold=True)
            
            ws[f'B{row}'] = description
            ws[f'B{row}'].border = self.thin_border
            ws[f'B{row}'].font = Font(name='Calibri', size=9)
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 50


    def _create_router_detail_sheet(self, wb: Workbook, router_mac: str, router_data: Dict):
        """Create detailed sheet for a specific router with robust null handling"""
        
        router_name = router_data['router_name']
        router_model = router_data['router_model']
        sheet_name = self._get_router_sheet_name(router_name, router_model)
        
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        ws = wb.create_sheet(sheet_name)
        
        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = f'🔧 {router_name}-{router_model} - Detailed Test History'
        ws['A1'].font = self.header_font
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], 
                                    end_color=self.colors['header'], 
                                    fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Router details
        row = 3
        total_duration_formatted = self._format_duration_from_seconds(router_data['total_duration'])
        
        details = [
            ('Router Name:', router_name),
            ('Model:', router_model),
            ('MAC Address:', router_data['router_mac']),
            ('Firmware:', router_data['router_firmware']),
            ('Total Tests:', router_data['total_tests']),
            ('Passed:', router_data['passed_tests']),
            ('Failed:', router_data['failed_tests']),
            ('Total Duration:', total_duration_formatted)
        ]
        
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.bold_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.normal_font
            row += 1
        
        # Back to summary link
        row += 1
        ws[f'A{row}'] = '← Back to Summary'
        ws[f'A{row}'].hyperlink = "#'Test Summary'!A1"
        ws[f'A{row}'].font = self.link_font
        
        # Feature-wise test details
        row += 3
        
        for feature_name, tests in sorted(router_data['features'].items()):
            # Feature header
            ws.merge_cells(f'A{row}:K{row}')
            ws[f'A{row}'] = f'Feature: {feature_name}'
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = PatternFill(start_color=self.colors['subheader'], 
                                            end_color=self.colors['subheader'], 
                                            fill_type='solid')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'A{row}'].border = self.thin_border
            row += 1
            
            # Test table headers
            headers = [
                'Date', 'Time', 'Scenario', 'Clients', 'Status', 
                'Router CPE\n(Creation %)', 'Router CPE\n(Test %)', 
                'Time Taken (s)', 'Test Duration', 'Steps', 'Failure Details'
            ]
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['header'], 
                                    end_color=self.colors['header'], 
                                    fill_type='solid')
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            ws.row_dimensions[row].height = 30
            row += 1
            
            # Sort tests by date (newest first)
            sorted_tests = sorted(tests, 
                                key=lambda x: x.get('test_time', ''), 
                                reverse=True)
            
            # Test data rows
            for test in sorted_tests:
                # Parse test time
                test_time = test.get('test_time', '')
                if test_time:
                    if isinstance(test_time, str):
                        try:
                            dt = datetime.fromisoformat(test_time.replace('Z', '+00:00'))
                            test_date = dt.strftime('%d %b %Y')
                            test_time_str = dt.strftime('%H:%M:%S')
                        except:
                            test_date = test_time[:10] if len(test_time) >= 10 else 'N/A'
                            test_time_str = 'N/A'
                    else:
                        test_date = test_time.strftime('%d %b %Y')
                        test_time_str = test_time.strftime('%H:%M:%S')
                else:
                    test_date = 'N/A'
                    test_time_str = 'N/A'
                
                status = test.get('status', 'unknown').lower()
                
                # Calculate test duration
                start_time = test.get('start_time')
                end_time = test.get('end_time')
                test_duration = self._calculate_duration(start_time, end_time) if start_time and end_time else '00:00:00'
                
                # Handle null values for clients
                clients_value = test.get('number_of_clients')
                clients_display = str(clients_value) if clients_value is not None else 'N/A'
                
                # Handle null values for CPU
                router_cpu_creation = test.get('router_avg_cpu_creation')
                router_cpu_test = test.get('router_avg_cpu_test')
                cpu_creation_display = round(router_cpu_creation, 2) if router_cpu_creation is not None else 'N/A'
                cpu_test_display = round(router_cpu_test, 2) if router_cpu_test is not None else 'N/A'
                
                # Handle null values for time taken
                time_taken = test.get('time_taken')
                time_taken_display = round(time_taken, 2) if time_taken is not None else 'N/A'
                
                # Get step summary
                steps_data = test.get('steps_data', [])
                if steps_data:
                    passed_steps = sum(1 for s in steps_data if s.get('status') == 'passed')
                    total_steps = len(steps_data)
                    failed_steps = sum(1 for s in steps_data if s.get('status') in ['failed', 'undefined'])
                    
                    if failed_steps > 0:
                        step_summary = f"❌ {passed_steps}/{total_steps}"
                    else:
                        step_summary = f"✅ {passed_steps}/{total_steps}"
                else:
                    step_summary = "N/A"
                
                # Get failure details
                failure_details = ""
                failure_type = ""
                if status != 'passed':
                    for step in steps_data:
                        step_status = step.get('status', '')
                        if step_status in ['failed', 'undefined']:
                            step_keyword = step.get('keyword', '')
                            step_name = step.get('name', '')
                            failure_msg = step.get('failure_message', '')
                            
                            # Categorize failure
                            if step_status == 'undefined':
                                failure_type = "🔧 UNDEFINED"
                            elif 'timeout' in failure_msg.lower():
                                failure_type = "⏱️ TIMEOUT"
                            elif 'assertion' in failure_msg.lower():
                                failure_type = "ASSERTION"
                            else:
                                failure_type = "ERROR"
                            
                            failure_details = f"{failure_type}\n{step_keyword} {step_name}\n{failure_msg}"
                            break
                    
                    if not failure_details:
                        failure_details = f"FAILED\n{test.get('failure_reason', 'Unknown error')}"
                
                # Data
                data_row = [
                    test_date,
                    test_time_str,
                    test.get('scenario_name', 'N/A'),
                    clients_display,
                    test.get('status', 'unknown').upper(),
                    cpu_creation_display,
                    cpu_test_display,
                    time_taken_display,
                    test_duration,
                    step_summary,
                    failure_details
                ]
                
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = value
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Calibri', size=9)
                    
                    # Color code status column
                    if col_idx == 5:
                        if status == 'passed':
                            cell.fill = PatternFill(start_color=self.colors['pass'], 
                                                end_color=self.colors['pass'], 
                                                fill_type='solid')
                            cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
                        else:
                            cell.fill = PatternFill(start_color=self.colors['fail'], 
                                                end_color=self.colors['fail'], 
                                                fill_type='solid')
                            cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
                    
                    # Highlight N/A values in gray
                    if value == 'N/A' and col_idx in [4, 6, 7, 8]:
                        cell.fill = PatternFill(start_color='E7E6E6', 
                                            end_color='E7E6E6', 
                                            fill_type='solid')
                        cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
                    
                    # Failure details - left align and wrap
                    if col_idx == 11:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        if failure_details:
                            # Color code based on failure type
                            if 'UNDEFINED' in failure_details:
                                cell.fill = PatternFill(start_color=self.colors['step_skipped'], 
                                                    end_color=self.colors['step_skipped'], 
                                                    fill_type='solid')
                            elif '⏱ TIMEOUT' in failure_details:
                                cell.fill = PatternFill(start_color='FFB366', 
                                                    end_color='FFB366', 
                                                    fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color=self.colors['step_failed'], 
                                                    end_color=self.colors['step_failed'], 
                                                    fill_type='solid')
                
                ws.row_dimensions[row].height = 30 if not failure_details else 60
                row += 1
            
            row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 55

    
    def update_statistics(self, all_tests: List[Dict]):
        """Update the Excel statistics file with all test data"""
        # Aggregate data by router
        aggregated_data = self._aggregate_router_data(all_tests)
        
        # Load or create workbook
        wb = load_workbook(self.excel_path)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        self._create_summary_sheet(wb, aggregated_data)
        
        
        # Create detail sheets for each router
        for router_mac, router_data in aggregated_data.items():
            self._create_router_detail_sheet(wb, router_mac, router_data)
        
        # Create failure analysis sheet
        self._create_failure_analysis_sheet(wb, aggregated_data)

        # Save workbook
        wb.save(self.excel_path)
        
        # Set file permissions
        try:
            os.chmod(self.excel_path, 0o666)
        except:
            pass
        
        return self.excel_path