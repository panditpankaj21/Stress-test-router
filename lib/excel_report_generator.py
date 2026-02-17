"""
Excel Report Generator for Test Results
Generates comprehensive Excel reports with conditional formatting
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import statistics
from collections import defaultdict


class ExcelReportGenerator:
    def __init__(self):
        """Initialize Excel report generator"""
        self.wb = Workbook()
        
        # Define color schemes
        self.colors = {
            'header': '366092',
            'good': '92D050',      # Green for good performance (CPU < 30%)
            'warning': 'FFC000',   # Yellow for moderate (30% <= CPU < 50%)
            'critical': 'FF0000',  # Red for critical (CPU >= 50%)
            'excellent': '00B050', # Dark green for excellent (CPU < 20%)
            'subheader': 'D9E1F2'
        }
        
        # Define styles
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Calibri', size=11, bold=True)
        self.normal_font = Font(name='Calibri', size=10)
        self.bold_font = Font(name='Calibri', size=10, bold=True)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _aggregate_test_data(self, all_tests: List[Dict]) -> Dict:
        """
        Aggregate test data by router, feature, and number of clients
        
        Returns structure:
        {
            'router_mac': {
                'router_name': 'Name',
                'router_model': 'Model',
                'features': {
                    'feature_name': {
                        'num_clients': {
                            'router_cpu_creation': [values],
                            'router_cpu_test': [values],
                            'linux_cpu_creation': [values],
                            'linux_cpu_test': [values],
                            'time_taken': [values],
                            'test_count': count
                        }
                    }
                }
            }
        }
        """
        aggregated = defaultdict(lambda: {
            'router_name': None,
            'router_model': None,
            'router_firmware': None,
            'features': defaultdict(lambda: defaultdict(lambda: {
                'router_cpu_creation': [],
                'router_cpu_test': [],
                'linux_cpu_creation': [],
                'linux_cpu_test': [],
                'time_taken': [],
                'test_count': 0
            }))
        })
        
        for test in all_tests:
            router_mac = test.get('router_mac')
            if not router_mac:
                continue
            
            # Set router info (from most recent test data)
            if not aggregated[router_mac]['router_name']:
                aggregated[router_mac]['router_name'] = test.get('router_name', 'Unknown')
                aggregated[router_mac]['router_model'] = test.get('router_model', 'Unknown')
                aggregated[router_mac]['router_firmware'] = test.get('router_firmware', 'Unknown')
            
            feature_name = test.get('feature_name', 'Unknown Feature')
            num_clients = test.get('number_of_clients', 0)
            
            # Aggregate metrics
            feature_data = aggregated[router_mac]['features'][feature_name][num_clients]
            
            if test.get('router_avg_cpu_creation') is not None:
                feature_data['router_cpu_creation'].append(test['router_avg_cpu_creation'])
            if test.get('router_avg_cpu_test') is not None:
                feature_data['router_cpu_test'].append(test['router_avg_cpu_test'])
            if test.get('linux_avg_cpu_creation') is not None:
                feature_data['linux_cpu_creation'].append(test['linux_avg_cpu_creation'])
            if test.get('linux_avg_cpu_test') is not None:
                feature_data['linux_cpu_test'].append(test['linux_avg_cpu_test'])
            if test.get('time_taken') is not None:
                feature_data['time_taken'].append(test['time_taken'])
            
            feature_data['test_count'] += 1
        
        return dict(aggregated)
    
    def _calculate_averages(self, values_list: List[float]) -> float:
        """Calculate average, handling empty lists"""
        if not values_list:
            return 0.0
        return statistics.mean(values_list)
    
    def _get_performance_color(self, cpu_value: float) -> str:
        """
        Determine color based on CPU utilization
        < 20%: Excellent (dark green)
        20-30%: Good (green)
        30-50%: Warning (yellow)
        >= 50%: Critical (red)
        """
        if cpu_value < 20:
            return self.colors['excellent']
        elif cpu_value < 30:
            return self.colors['good']
        elif cpu_value < 50:
            return self.colors['warning']
        else:
            return self.colors['critical']
    
    def _apply_cell_style(self, cell, value, is_header=False, is_subheader=False, 
                         apply_cpu_color=False, is_bold=False):
        """Apply styling to a cell"""
        cell.value = value
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if is_header:
            cell.font = self.header_font
            cell.fill = PatternFill(start_color=self.colors['header'], 
                                   end_color=self.colors['header'], 
                                   fill_type='solid')
        elif is_subheader:
            cell.font = self.subheader_font
            cell.fill = PatternFill(start_color=self.colors['subheader'], 
                                   end_color=self.colors['subheader'], 
                                   fill_type='solid')
        elif apply_cpu_color and isinstance(value, (int, float)):
            cell.font = self.bold_font if value >= 50 else self.normal_font
            color = self._get_performance_color(value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        elif is_bold:
            cell.font = self.bold_font
        else:
            cell.font = self.normal_font
    
    def _create_summary_sheet(self, ws, aggregated_data: Dict):
        """Create summary sheet with overall statistics"""
        ws.title = "Summary"
        
        # Title
        ws.merge_cells('A1:F1')
        self._apply_cell_style(ws['A1'], '📊 Test Execution Summary Report', is_header=True)
        ws.row_dimensions[1].height = 30
        
        # Report metadata
        row = 3
        ws[f'A{row}'] = 'Report Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self._apply_cell_style(ws[f'A{row}'], ws[f'A{row}'].value, is_bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Total Routers Tested:'
        ws[f'B{row}'] = len(aggregated_data)
        self._apply_cell_style(ws[f'A{row}'], ws[f'A{row}'].value, is_bold=True)
        
        # Summary table header
        row += 3
        headers = ['Router Name (Model)', 'MAC Address', 'Firmware', 'Features Tested', 
                  'Total Tests', 'Avg Router CPU %', 'Avg Linux CPU %']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            self._apply_cell_style(cell, header, is_header=True)
        
        # Summary data
        row += 1
        for router_mac, router_data in aggregated_data.items():
            router_name = router_data['router_name']
            router_model = router_data['router_model']
            router_firmware = router_data['router_firmware']
            
            # Calculate overall statistics
            all_router_cpu = []
            all_linux_cpu = []
            total_tests = 0
            num_features = len(router_data['features'])
            
            for feature_name, client_data in router_data['features'].items():
                for num_clients, metrics in client_data.items():
                    all_router_cpu.extend(metrics['router_cpu_test'])
                    all_linux_cpu.extend(metrics['linux_cpu_test'])
                    total_tests += metrics['test_count']
            
            avg_router_cpu = self._calculate_averages(all_router_cpu)
            avg_linux_cpu = self._calculate_averages(all_linux_cpu)
            
            # Write row
            ws.cell(row=row, column=1).value = f"{router_name} ({router_model})"
            ws.cell(row=row, column=2).value = router_mac
            ws.cell(row=row, column=3).value = router_firmware
            ws.cell(row=row, column=4).value = num_features
            ws.cell(row=row, column=5).value = total_tests
            
            # Apply CPU coloring
            self._apply_cell_style(ws.cell(row=row, column=6), 
                                 round(avg_router_cpu, 2), 
                                 apply_cpu_color=True)
            self._apply_cell_style(ws.cell(row=row, column=7), 
                                 round(avg_linux_cpu, 2), 
                                 apply_cpu_color=True)
            
            # Apply border to other cells
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', 
                                                                   vertical='center')
            
            row += 1
        
        # Add legend
        row += 2
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'Performance Color Legend:'
        self._apply_cell_style(ws[f'A{row}'], ws[f'A{row}'].value, is_bold=True)
        
        row += 1
        legend_items = [
            ('Excellent (< 20%)', self.colors['excellent']),
            ('Good (20-30%)', self.colors['good']),
            ('Warning (30-50%)', self.colors['warning']),
            ('Critical (≥ 50%)', self.colors['critical'])
        ]
        
        for idx, (label, color) in enumerate(legend_items, start=1):
            cell = ws.cell(row=row, column=idx)
            cell.value = label
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = self.bold_font
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
    
    def _create_router_sheet(self, router_mac: str, router_data: Dict):
        """Create detailed sheet for each router"""
        router_name = router_data['router_name']
        router_model = router_data['router_model']
        
        # Create sheet with safe name (Excel sheet names have limitations)
        sheet_name = f"{router_name[:25]}-{router_model}"
        ws = self.wb.create_sheet(title=sheet_name)
        
        # Title
        ws.merge_cells('A1:K1')
        title = f"🔧 {router_name} ({router_model})"
        self._apply_cell_style(ws['A1'], title, is_header=True)
        ws.row_dimensions[1].height = 30
        
        # Router info
        row = 3
        ws[f'A{row}'] = 'MAC Address:'
        ws[f'B{row}'] = router_mac
        self._apply_cell_style(ws[f'A{row}'], ws[f'A{row}'].value, is_bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Firmware:'
        ws[f'B{row}'] = router_data['router_firmware']
        self._apply_cell_style(ws[f'A{row}'], ws[f'A{row}'].value, is_bold=True)
        
        # Data table
        row += 3
        headers = [
            'Feature/Test', 'Clients', 'Tests Run',
            'Avg Router CPU\n(Creation) %', 'Avg Router CPU\n(Test) %',
            'Avg Linux CPU\n(Creation) %', 'Avg Linux CPU\n(Test) %',
            'Avg Time Taken (s)',
            'Max Router CPU %', 'Max Linux CPU %', 'Status'
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            self._apply_cell_style(cell, header, is_header=True)
            ws.row_dimensions[row].height = 30
        
        # Data rows
        row += 1
        for feature_name, client_data in sorted(router_data['features'].items()):
            # Feature header row
            feature_row = row
            ws.merge_cells(f'A{row}:K{row}')
            self._apply_cell_style(ws[f'A{row}'], f"📁 {feature_name}", is_subheader=True)
            row += 1
            
            # Sort by number of clients
            for num_clients in sorted(client_data.keys()):
                metrics = client_data[num_clients]
                
                # Calculate averages
                avg_router_cpu_creation = self._calculate_averages(metrics['router_cpu_creation'])
                avg_router_cpu_test = self._calculate_averages(metrics['router_cpu_test'])
                avg_linux_cpu_creation = self._calculate_averages(metrics['linux_cpu_creation'])
                avg_linux_cpu_test = self._calculate_averages(metrics['linux_cpu_test'])
                avg_time_taken = self._calculate_averages(metrics['time_taken'])
                
                max_router_cpu = max(metrics['router_cpu_test']) if metrics['router_cpu_test'] else 0
                max_linux_cpu = max(metrics['linux_cpu_test']) if metrics['linux_cpu_test'] else 0
                
                # Determine status
                if max_router_cpu >= 50 or max_linux_cpu >= 50:
                    status = '⚠️ Critical'
                elif avg_router_cpu_test >= 50 or avg_linux_cpu_test >= 50:
                    status = '⚠️ Warning'
                elif avg_router_cpu_test < 20 and avg_linux_cpu_test < 20:
                    status = '✅ Excellent'
                else:
                    status = '✓ Good'
                
                # Write data
                col = 1
                ws.cell(row=row, column=col).value = f"  → {num_clients} clients"
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=row, column=col).border = self.thin_border
                
                col += 1
                ws.cell(row=row, column=col).value = num_clients
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                
                col += 1
                ws.cell(row=row, column=col).value = metrics['test_count']
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                
                # CPU columns with coloring
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(avg_router_cpu_creation, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(avg_router_cpu_test, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(avg_linux_cpu_creation, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(avg_linux_cpu_test, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                ws.cell(row=row, column=col).value = round(avg_time_taken, 2)
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(max_router_cpu, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                self._apply_cell_style(ws.cell(row=row, column=col), 
                                     round(max_linux_cpu, 2), 
                                     apply_cpu_color=True)
                
                col += 1
                ws.cell(row=row, column=col).value = status
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=row, column=col).font = self.bold_font
                
                row += 1
            
            row += 1  # Space between features
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 16
        ws.column_dimensions['K'].width = 15
    
    def generate_excel_report(self, all_tests: List[Dict], output_path: str) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            all_tests: List of all test documents from MongoDB
            output_path: Path where to save the Excel file
            
        Returns:
            Path to generated Excel file
        """
        # Aggregate data
        aggregated_data = self._aggregate_test_data(all_tests)
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Create summary sheet
        ws_summary = self.wb.create_sheet(title="Summary", index=0)
        self._create_summary_sheet(ws_summary, aggregated_data)
        
        # Create individual router sheets
        for router_mac, router_data in aggregated_data.items():
            self._create_router_sheet(router_mac, router_data)
        
        # Save workbook
        self.wb.save(output_path)
        return output_path